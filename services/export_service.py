"""
Export Service
Handles exporting transactions to Excel and CSV formats with accounting-style display
"""
import csv
from pathlib import Path
from typing import List, Dict, Optional
from datetime import datetime, date
from decimal import Decimal
from tkinter import filedialog

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from models import Transaction
from config import Config


class ExportService:
    """Service for exporting transactions to various formats"""

    def __init__(self):
        """Initialize export service"""
        self.export_dir = Config.EXPORT_DIR
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def export_to_excel(
        self,
        transactions: List[Transaction],
        filename: str = None,
        start_date: date = None,
        end_date: date = None
    ) -> Optional[Path]:
        """
        Export transactions to Excel file with two sheets

        Args:
            transactions: List of transactions to export
            filename: Optional custom filename (without extension)
            start_date: Optional start date for period label
            end_date: Optional end date for period label

        Returns:
            Path to created Excel file, or None if user cancelled

        Raises:
            ImportError: If openpyxl is not installed
            Exception: If export fails
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )

        # Generate default filename if not provided
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"transactions_{timestamp}"

        # Ensure .xlsx extension
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'

        # Ask user to choose save location
        filepath = filedialog.asksaveasfilename(
            title="Save Excel Export",
            initialdir=str(self.export_dir),
            initialfile=filename,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )

        # User cancelled
        if not filepath:
            return None

        filepath = Path(filepath)

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create Transactions sheet
        self._create_transactions_sheet(wb, transactions)

        # Create Summary sheet
        self._create_summary_sheet(wb, transactions, start_date, end_date)

        # Save workbook
        wb.save(filepath)

        return filepath

    def _create_transactions_sheet(self, wb: Workbook, transactions: List[Transaction]):
        """Create the Transactions sheet with accounting columns"""
        ws = wb.create_sheet("Transactions", 0)

        # Define headers
        headers = ["Date", "Description", "Account", "Category", "Type", "Debit", "Credit", "Notes"]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write transaction data
        for row_num, transaction in enumerate(transactions, 2):
            formatted_data = self._format_for_accounting(transaction)

            ws.cell(row=row_num, column=1, value=formatted_data['Date'])
            ws.cell(row=row_num, column=2, value=formatted_data['Description'])
            ws.cell(row=row_num, column=3, value=formatted_data['Account'])
            ws.cell(row=row_num, column=4, value=formatted_data['Category'])
            ws.cell(row=row_num, column=5, value=formatted_data['Type'])

            # Debit column
            debit_cell = ws.cell(row=row_num, column=6, value=formatted_data['Debit'])
            if formatted_data['Debit'] != '-':
                debit_cell.number_format = '#,##0.00'
                debit_cell.font = Font(color="C00000")  # Red for debits

            # Credit column
            credit_cell = ws.cell(row=row_num, column=7, value=formatted_data['Credit'])
            if formatted_data['Credit'] != '-':
                credit_cell.number_format = '#,##0.00'
                credit_cell.font = Font(color="00B050")  # Green for credits

            ws.cell(row=row_num, column=8, value=formatted_data['Notes'])

        # Auto-adjust column widths
        column_widths = {
            'A': 12,  # Date
            'B': 25,  # Description
            'C': 20,  # Account
            'D': 20,  # Category
            'E': 10,  # Type
            'F': 15,  # Debit
            'G': 15,  # Credit
            'H': 30   # Notes
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

    def _create_summary_sheet(
        self,
        wb: Workbook,
        transactions: List[Transaction],
        start_date: date = None,
        end_date: date = None
    ):
        """Create the Summary sheet with totals and statistics"""
        ws = wb.create_sheet("Summary", 1)

        # Calculate summary statistics
        summary = self._calculate_summary(transactions)

        # Title
        ws['A1'] = "Transaction Summary"
        ws['A1'].font = Font(bold=True, size=14)

        # Period
        row = 3
        if start_date and end_date:
            ws[f'A{row}'] = "Period:"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = f"{start_date.strftime('%d-%m-%Y')} to {end_date.strftime('%d-%m-%Y')}"
            row += 1
        elif start_date:
            ws[f'A{row}'] = "From Date:"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = start_date.strftime('%d-%m-%Y')
            row += 1
        elif end_date:
            ws[f'A{row}'] = "To Date:"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = end_date.strftime('%d-%m-%Y')
            row += 1

        row += 1

        # Summary data
        summary_data = [
            ("Total Transactions:", summary['total_count']),
            ("", ""),
            ("Total Credits:", f"Rs. {summary['total_credits']:,.2f}"),
            ("Total Debits:", f"Rs. {summary['total_debits']:,.2f}"),
            ("", ""),
            ("Net Balance:", f"Rs. {summary['net_balance']:,.2f}"),
            ("", ""),
            ("Credit Transactions:", summary['credit_count']),
            ("Debit Transactions:", summary['debit_count']),
        ]

        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value

            # Color code net balance
            if label == "Net Balance:":
                if summary['net_balance'] > 0:
                    ws[f'B{row}'].font = Font(bold=True, color="00B050")  # Green
                elif summary['net_balance'] < 0:
                    ws[f'B{row}'].font = Font(bold=True, color="C00000")  # Red

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def export_to_csv(
        self,
        transactions: List[Transaction],
        filename: str = None
    ) -> Optional[Path]:
        """
        Export transactions to CSV file

        Args:
            transactions: List of transactions to export
            filename: Optional custom filename (without extension)

        Returns:
            Path to created CSV file, or None if user cancelled
        """
        # Generate default filename if not provided
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"transactions_{timestamp}"

        # Ensure .csv extension
        if not filename.endswith('.csv'):
            filename += '.csv'

        # Ask user to choose save location
        filepath = filedialog.asksaveasfilename(
            title="Save CSV Export",
            initialdir=str(self.export_dir),
            initialfile=filename,
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )

        # User cancelled
        if not filepath:
            return None

        filepath = Path(filepath)

        # Write CSV
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ["Date", "Description", "Account", "Category", "Type", "Debit", "Credit", "Notes"]
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

            writer.writeheader()

            for transaction in transactions:
                formatted_data = self._format_for_accounting(transaction)
                writer.writerow(formatted_data)

        return filepath

    def _format_for_accounting(self, transaction: Transaction) -> Dict:
        """
        Format transaction data for accounting-style display

        Args:
            transaction: Transaction object

        Returns:
            Dictionary with formatted data including Debit/Credit columns
        """
        # Determine debit and credit values
        if transaction.transaction_type == 'credit':
            debit = '-'
            credit = float(transaction.amount)
        elif transaction.transaction_type == 'debit':
            debit = float(transaction.amount)
            credit = '-'
        else:  # transfer or other
            # For transfers, show as debit
            debit = float(transaction.amount)
            credit = '-'

        return {
            'Date': transaction.date.strftime('%d-%m-%Y'),
            'Description': transaction.description or '',
            'Account': transaction.account.name if transaction.account else '',
            'Category': transaction.category.name if transaction.category else 'Uncategorized',
            'Type': transaction.transaction_type.capitalize() if transaction.transaction_type else '',
            'Debit': debit,
            'Credit': credit,
            'Notes': transaction.notes or ''
        }

    def _calculate_summary(self, transactions: List[Transaction]) -> Dict:
        """
        Calculate summary statistics for transactions

        Args:
            transactions: List of transactions

        Returns:
            Dictionary with summary statistics
        """
        total_credits = Decimal('0')
        total_debits = Decimal('0')
        credit_count = 0
        debit_count = 0

        for transaction in transactions:
            if transaction.transaction_type == 'credit':
                total_credits += Decimal(str(transaction.amount))
                credit_count += 1
            elif transaction.transaction_type == 'debit':
                total_debits += Decimal(str(transaction.amount))
                debit_count += 1

        net_balance = total_credits - total_debits

        return {
            'total_count': len(transactions),
            'total_credits': float(total_credits),
            'total_debits': float(total_debits),
            'net_balance': float(net_balance),
            'credit_count': credit_count,
            'debit_count': debit_count
        }
